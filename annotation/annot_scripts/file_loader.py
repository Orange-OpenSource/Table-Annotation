"""
* Software Name : TableAnnotation
* Author: Viet-Phi Huynh, Jixiong Liu, Yoan Chabot, Frédéric Deuzé and Raphael Troncy
* Software description: TableAnnotation (a.k.a DAGOBAH) is a semantic annotation tool for tables leveraging three steps: 1) Table Preprocessing: a set of comprehensive heuristic to clean the table (e.g. fix encoding error), determine table orientation, data types of columns. 2) Entity Lookup: retrieve a number of entity candidates for mentions in the table, using an elastic search-based entity lookup. 3) Annotation: disambiguate retrieved entity candidates, select the most relevant entity for each mention. This consists of three tasks, namely Cell-Entity Annotation, Column-Type Annotation, Column-Pair Annotation.
* Version: <1.0.0>
* SPDX-FileCopyrightText: Copyright (c) 2023 Orange
* SPDX-License-Identifier: GPL-3.0-or-later
* Licensed under the GNU-GPL v3 (the "License");
 * you may not use this file except in compliance with the License.
 * You may obtain a copy of the License at
 *
 * https://www.gnu.org/licenses/gpl-3.0.html#license-text
 *
 * Unless required by applicable law or agreed to in writing, software
 * distributed under the License is distributed on an "AS IS" BASIS,
 * WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
 * See the License for the specific language governing permissions and
 * limitations under the License.
"""
'''
FILE READER:

Auto converter

Auto-converter is for automatically transform from different types of data
to tables

'''
import csv
import pandas as pd
import numpy as np
import chardet
import datetime
import openpyxl
from openpyxl.utils import range_boundaries
from scipy import ndimage as ndi

def txt_to_table(filepath: str):
    """
    Read table from text file (txt, tsv,csv..). Currently, only 1 table per file supported 
                                        and delimiter detected automatically.
    """
    list_tables = []
    try:
        f =  open(filepath, 'rb')
        en_scheme = chardet.detect(f.read())  # detect encoding scheme
        f.close()
        f = open(filepath, 'r', encoding=en_scheme['encoding'])
        possible_sep = [',', '\t', ';', ':']
        dialect = csv.Sniffer().sniff(f.read(), possible_sep)
        f.close()
        f = open(filepath, 'r', encoding=en_scheme['encoding'])
    except FileNotFoundError:
        print(" File not found !!")
        return []
    except Exception as e:
        print(e)
        return []
    reader = csv.reader(f, delimiter=dialect.delimiter, skipinitialspace=True)
    table = []
    for item in reader:
        table.append(item)
    if table:
        list_tables.append(table)
    f.close()
    return {"tableFromTextFile": list_tables}

"""
def json_to_table(filepath: str) -> List[List[List[str]]]:
    list_tables = []
    table = []
    with open(filepath, 'r', encoding="utf8") as f:
        temp = json.loads(f.read())
        table.append(temp)
    list_tables.append(table)
    return list_tables
"""

def excel_to_table(filepath):
    """
    Read multiple tables per worksheet in excel file. Only .xlsx supported. Old .xls not supported.
    """
    wb_obj = openpyxl.load_workbook(filepath, data_only=True)
    sheetnames = wb_obj.sheetnames
    tables_per_sheet = {}
    for sheet_name in sheetnames:
        w_sheet = wb_obj[sheet_name]
        num_sheet_col = w_sheet.max_column
        num_sheet_row = w_sheet.max_row
        ## unmerge cells
        groups = w_sheet.merged_cells
        for group in list(groups):
            w_sheet.unmerge_cells(range_string=str(group))
            min_col, min_row, max_col, max_row = range_boundaries(str(group))
            top_left_cell_value = w_sheet.cell(row=min_row, column=min_col).value
            for row in w_sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value

        ## clustering tables in a worksheet by connected components.
        raw_sheet = [] ## read value of all cells.
        binary_sheet = []  ## 1 (foreground) if cell contains value, otherwise 0 (background)
        for row in w_sheet.iter_rows(min_col=1, min_row=1, max_col=num_sheet_col, max_row=num_sheet_row):
            binary_row = []
            value_row = []
            for cell in row:
                ## read raw value, take care of reading datetime properly.
                if cell.value:
                    if isinstance(cell.value, datetime.datetime):
                        value_row.append(cell.value.strftime('%m/%d/%Y'))
                    else:
                        value_row.append(cell.value)
                else:
                    value_row.append("")
                    
                ## decide if a cell contains a value (hence a foreground cell)
                if cell.value:
                    binary_row.append(True)
                elif cell.fill.patternType:
                    binary_row.append(True)
                elif cell.border.left.style or cell.border.right.style:
                    binary_row.append(True)
                else:
                    binary_row.append(False)
            raw_sheet.append(value_row)
            binary_sheet.append(binary_row)

        ## find connected components
        cnt_components = np.array(binary_sheet)
        cnt_component_labels, n_cnt_component = ndi.label(cnt_components)

        ## each connected component can be a potential independent table.
        tables = []
        for i_label in range(1, n_cnt_component + 1):
            ## define the rectangle that may contain a table
            min_row = 0
            max_row = 0
            min_col = 0
            max_col = 0
            for i_line, line in enumerate(cnt_component_labels):
                if i_label in line:
                    min_row = i_line
                    break
            for i_line, line in enumerate(list(map(list, zip(*cnt_component_labels)))):
                if i_label in line:
                    min_col = i_line
                    break
            for i_line, line in enumerate(cnt_component_labels[::-1]):
                if i_label in line:
                    max_row = len(cnt_component_labels) - 1 - i_line
                    break
            for i_line, line in enumerate(list(map(list, zip(*cnt_component_labels)))[::-1]):
                if i_label in line:
                    max_col = num_sheet_col - 1 - i_line
                    break 
            ## check if there exist potentially a table in the rectangle.
            table = np.array(raw_sheet)[min_row:max_row+1,min_col:max_col+1]
            if table.shape[0] > 1 and table.shape[1] > 1:
                tables.append(table.tolist())
        tables_per_sheet[f"tableFromExcelSheet_{sheet_name}"] = tables
    return tables_per_sheet

def deprecated_excel_to_table(filepath):
    """
    Read table from excel file. Currently, multi tables supported.
    Only one heuristic supported for seperating tables: blank lines between 2 consecutive tables.
    Args: 
        input table path
    Return
        3D array with first dimension is the number of 2D tables.
    """
    list_tables = []
    try:
        xl = pd.ExcelFile(filepath, engine='openpyxl')
    except FileNotFoundError:
        print(" File not found !!")
        return []
    except:
        return []
    sheet_names = xl.sheet_names
    for i_sheet in range(len(sheet_names)):
        excel_sheet = pd.read_excel(filepath, header=None, sheet_name=i_sheet)
        excel_sheet = excel_sheet.values.tolist()  
        single_table = []
        for line in excel_sheet:
            i_element = len(line) - 1
            end_of_line = len(line) - 1
            is_eol = False
            while i_element >= 0:
                if pd.isna(line[i_element]):
                    if not is_eol:
                        end_of_line -= 1
                    else:
                        line[i_element] = ""
                else:
                    is_eol = True
                i_element -= 1
            line = line[:end_of_line+1]
            line = [str(s) for s in line]
            if line == []:
                if single_table != []:
                    list_tables.append(single_table)
                    single_table = []
            else: 
                tmp_line = []
                for e in line:
                    if e != "" and e != " "*len(e):
                        tmp_line.append(e)       
                if tmp_line == []:
                    if len(single_table) > 1:
                        single_table.append(line)
                    else:
                        single_table = []
                else:
                    single_table.append(line)       
        if single_table != []:
            list_tables.append(single_table)
    return list_tables

def file_loader(filepath):
    '''
    automatic form detection tool.
    :param filepath: target file
    :return:
    '''
    splitPart = filepath.split('.')
    if(splitPart[-1].lower() in ['csv', 'txt', 'tsv']): return txt_to_table(filepath)
    # if(splitPart[-1] in ['json']): return json_to_table(filepath)
    if(splitPart[-1].lower() in ['xlsx']): return excel_to_table(filepath)
    return []


if __name__ == '__main__':
    # csv = file_loader(r"C:\Users\pgkx5469\Documents\ECE.csv")
    txt = file_loader(r"/datastorage/uploaded_files/ECE.csv")
    # txt = file_loader(r"C:\Users\pgkx5469\Documents\Python Scripts\t1.txt")
    # excel = file_loader(r"C:\Users\pgkx5469\Documents\Projets\dagobah\data\round_3\1.xlsx")
    print(txt)